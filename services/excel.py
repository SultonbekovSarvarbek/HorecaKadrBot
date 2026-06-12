"""Выгрузка кандидатов в Excel (.xlsx)."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import texts
from db.models import Candidate

HEADERS = [
    "ID",
    "Имя",
    "Возраст",
    "Телефон",
    "Район",
    "Вакансия",
    "Опыт",
    "График 6/1 или 5/2",
    "Готов выйти",
    "Статус",
    "Источник",
    "Username",
    "Telegram ID",
    "Дата заявки",
]


def export_candidates_xlsx(candidates: list[Candidate]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Кандидаты"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for c in candidates:
        ws.append(
            [
                c.id,
                c.name,
                c.age,
                c.phone,
                c.district,
                texts.VACANCY_LABELS[c.vacancy],
                c.experience,
                "Да" if c.schedule_ok else "Нет",
                c.start_when,
                texts.STATUS_LABELS[c.status],
                c.source or "",
                c.username or "",
                c.tg_id,
                c.created_at.strftime("%d.%m.%Y %H:%M"),
            ]
        )

    # автоширина колонок
    for col_idx in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
