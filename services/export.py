"""Экспорт анкет в Excel.

Фаза 2: сюда же добавится запись в Google Sheets API — интерфейс
export_candidates_xlsx(candidates) -> bytes останется без изменений,
добавится export_to_sheets(candidates) с той же подготовкой строк (см. _rows).
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import texts
from db.models import Candidate

HEADERS = [
    "Дата",
    "ФИО",
    "Возраст",
    "Пол",
    "Телефон",
    "Username",
    "Вакансия",
    "Филиал",
    "Район",
    "Опыт",
    "Русский",
    "Военный билет",
    "Свинина/алкоголь",
    "Источник",
    "Статус",
    "Причина отказа",
]


def _experience(c: Candidate) -> str:
    if c.cook_years is not None:
        spec = texts.COOK_SPEC_LABELS.get(c.cook_spec, "") if c.cook_spec else ""
        return f"{c.cook_years} лет {spec}".strip()
    return c.experience_cat or ""


def _reason(c: Candidate) -> str:
    parts = []
    if c.rejection_reason:
        parts.append(texts.REJECTION_LABELS.get(c.rejection_reason, c.rejection_reason.value))
    if c.status_reason:
        parts.append(c.status_reason)
    return "; ".join(parts)


def _rows(candidates: list[Candidate]) -> list[list]:
    rows = []
    for c in candidates:
        military = ""
        if c.military_id is not None:
            military = "Да" if c.military_id else "Нет"
        rows.append(
            [
                c.created_at.strftime("%d.%m.%Y %H:%M"),
                c.full_name,
                c.age,
                texts.GENDER_LABELS.get(c.gender, ""),
                c.phone,
                c.username or "",
                texts.POSITION_LABELS.get(c.vacancy.position, ""),
                c.vacancy.branch.name,
                c.district,
                _experience(c),
                texts.RUSSIAN_LABELS.get(c.russian, ""),
                military,
                "Да" if c.pork_alcohol_ok else "Нет",
                c.source or "",
                texts.STATUS_LABELS.get(c.status, c.status.value),
                _reason(c),
            ]
        )
    return rows


def export_candidates_xlsx(candidates: list[Candidate]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Анкеты"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in _rows(candidates):
        ws.append(row)
    for col_idx in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 40)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
